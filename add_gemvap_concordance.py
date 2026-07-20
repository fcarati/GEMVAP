"""
Add three colour-only concordance columns ("GEMVAP 1", "GEMVAP 2", "GEMVAP
3" — same naming convention as the existing "REVEL" / "AlphaMissense"
concordance columns) that compare each GEMVAP {1,2,3} CLASS column against
EXPERT CLASSIFICATION, using the identical static fill colours already used
by the REVEL / AlphaMissense concordance columns:
    green  FF00B050 - GEMVAP class matches expert tier
    yellow FFFFC000 - mismatch, but GEMVAP or expert tier is "ambiguous"
    red    FFFF0000 - mismatch, neither tier is "ambiguous"
No cell value is written, only the fill (matching how REVEL/AlphaMissense
concordance columns work: text-free, colour-only).
"""
import openpyxl

XLSX_PATH = r"c:\Users\Admin\Desktop\PhD\VSC_2\gemvap_clean_pipeline\data\raw\expert_missense.xlsx"
SHEET = "Expert_missenses_91"
DATA_LAST_ROW = 92

CLASS_COL_HEADER = {
    "GEMVAP_1": "GEMVAP 1 CLASS",
    "GEMVAP_2": "GEMVAP 2 CLASS",
    "GEMVAP_3": "GEMVAP 3 CLASS",
}
NEW_COL_HEADER = {
    "GEMVAP_1": "GEMVAP 1",
    "GEMVAP_2": "GEMVAP 2",
    "GEMVAP_3": "GEMVAP 3",
}
MODELS = ["GEMVAP_1", "GEMVAP_2", "GEMVAP_3"]

GREEN = "FF00B050"
YELLOW = "FFFFC000"
RED = "FFFF0000"

EXPERT_TIER = {
    "Pathogenic": "pathogenic",
    "Likely pathogenic": "pathogenic",
    "Benign": "benign",
    "Likely benign": "benign",
    "Uncertain significance": "ambiguous",
}


def concordance_color(predicted_tier, expert_tier):
    if predicted_tier is None or expert_tier is None:
        return None
    if predicted_tier == expert_tier:
        return GREEN
    if predicted_tier == "ambiguous" or expert_tier == "ambiguous":
        return YELLOW
    return RED


def main():
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb[SHEET]

    header = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    expert_col = header["EXPERT CLASSIFICATION"]

    unmapped = set()
    start_col = ws.max_column + 1
    for offset, model_name in enumerate(MODELS):
        class_col = header[CLASS_COL_HEADER[model_name]]
        new_col = start_col + offset
        ws.cell(row=1, column=new_col, value=NEW_COL_HEADER[model_name])

        counts = {"GREEN": 0, "YELLOW": 0, "RED": 0, "NONE": 0}
        for r in range(2, DATA_LAST_ROW + 1):
            predicted_tier = ws.cell(row=r, column=class_col).value
            expert_raw = ws.cell(row=r, column=expert_col).value
            expert_tier = EXPERT_TIER.get(expert_raw)
            if expert_raw is not None and expert_tier is None:
                unmapped.add(expert_raw)

            color = concordance_color(predicted_tier, expert_tier)
            cell = ws.cell(row=r, column=new_col)
            if color:
                cell.fill = openpyxl.styles.PatternFill(start_color=color, end_color=color, fill_type="solid")
                counts[{"FF00B050": "GREEN", "FFFFC000": "YELLOW", "FFFF0000": "RED"}[color]] += 1
            else:
                counts["NONE"] += 1
        print(model_name, counts)

    if unmapped:
        print(f"WARNING: unmapped EXPERT CLASSIFICATION values: {unmapped}")

    wb.save(XLSX_PATH)
    print(f"Saved {XLSX_PATH}")


if __name__ == "__main__":
    main()
