"""
Score the 91 expert_missense.xlsx FBN1 variants with the CURRENTLY FITTED
GEMVAP_1/2/3 models and classify them into benign/ambiguous/pathogenic using
the same PP3/BP4 Pejaver evidence thresholds the pipeline already calibrated
(output/gemvap_notebook_gnomad4/calibration/GEMVAP_{1,2,3}_thresholds.csv).

Appends three new columns ("GEMVAP 1 CLASS", "GEMVAP 2 CLASS", "GEMVAP 3
CLASS") after the last existing column, and extends the workbook's existing
REVEL-style conditional-formatting rule (text-based fill on
"pathogenic"/"benign"/"ambiguous", the same rule that colours "REVEL P") to
cover them, so the new columns get the identical colour code as REVEL P.
"""
import sys
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))

from gemvap_pipeline.model import load_fit_result, apply_consensus_score
from gemvap_pejaver_calibration import annotate_variant

XLSX_PATH = HERE / "data" / "raw" / "expert_missense.xlsx"
SHEET = "Expert_missenses_91"
RAW_TSV = HERE / "data" / "raw" / "FBN1_tableS1_allmissense_gnomad4.tsv"
OUTPUT_DIR = HERE / "output" / "gemvap_notebook_gnomad4"
CALIB_DIR = OUTPUT_DIR / "calibration"
DATA_LAST_ROW = 92  # header is row 1; 91 variants in rows 2-92; rows 95-96 are footer notes

MODELS = ["GEMVAP_1", "GEMVAP_2", "GEMVAP_3"]
NEW_COL_HEADER = {
    "GEMVAP_1": "GEMVAP 1 CLASS",
    "GEMVAP_2": "GEMVAP 2 CLASS",
    "GEMVAP_3": "GEMVAP 3 CLASS",
}


def evidence_to_tier(evidence: str) -> str:
    if evidence.startswith("PP3"):
        return "pathogenic"
    if evidence.startswith("BP4"):
        return "benign"
    return "ambiguous"  # Indeterminate


def load_thresholds(model_name: str) -> dict:
    thr = pd.read_csv(CALIB_DIR / f"{model_name}_thresholds.csv")
    return dict(zip(thr["evidence_level"], thr["score_threshold"]))


def find_revel_p_cf(ws):
    """Locate the ConditionalFormatting group that colours the REVEL P column
    (the containsText rules for pathogenic/benign/ambiguous/PP3/BP4)."""
    for cf in ws.conditional_formatting:
        rule_types = {r.dxfId for r in cf.rules}
        if rule_types == {1, 2, 3, 4, 5}:
            return cf
    raise RuntimeError("Could not find the REVEL P text-based conditional formatting group")


def main():
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb[SHEET]

    header = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    cdna_col = header["cDNA"]

    cdnas = [ws.cell(row=r, column=cdna_col).value for r in range(2, DATA_LAST_ROW + 1)]
    cdnas_clean = [c.strip() if isinstance(c, str) else None for c in cdnas]
    valid_cdnas = [c for c in cdnas_clean if c and c.startswith("c.")]

    print(f"Rows: {DATA_LAST_ROW - 1}, variants with a cDNA value: {len(valid_cdnas)}")

    raw = pd.read_csv(RAW_TSV, sep="\t", low_memory=False)
    raw = raw.drop_duplicates(subset="cDNA", keep="first").set_index("cDNA")

    missing = [c for c in valid_cdnas if c not in raw.index]
    if missing:
        print(f"WARNING: {len(missing)} cDNA values not found in {RAW_TSV.name}: {missing}")

    matched = raw.reindex(valid_cdnas)

    tiers_by_model = {}
    for model_name in MODELS:
        fit = load_fit_result(OUTPUT_DIR / f"{model_name}_fit.pkl")
        top_predictors = fit["top_predictors"]
        case_thresholds = fit["rbc"]["threshold"]["case"]

        scores = apply_consensus_score(matched, top_predictors, case_thresholds)
        pejaver_thresholds = load_thresholds(model_name)

        tiers = {}
        for cdna, score in zip(valid_cdnas, scores):
            if cdna not in raw.index or pd.isna(score):
                tiers[cdna] = None
                continue
            evidence = annotate_variant(float(score), pejaver_thresholds)
            tiers[cdna] = evidence_to_tier(evidence)
        tiers_by_model[model_name] = tiers

    # --- append new columns after the last existing column ---
    start_col = ws.max_column + 1
    for offset, model_name in enumerate(MODELS):
        col_idx = start_col + offset
        ws.cell(row=1, column=col_idx, value=NEW_COL_HEADER[model_name])
        for i, cdna in enumerate(cdnas_clean):
            r = i + 2
            tier = tiers_by_model[model_name].get(cdna) if cdna else None
            ws.cell(row=r, column=col_idx, value=tier)

    end_col_letter = get_column_letter(start_col + len(MODELS) - 1)
    start_col_letter = get_column_letter(start_col)
    new_range = f"{start_col_letter}2:{end_col_letter}{DATA_LAST_ROW}"

    revel_cf = find_revel_p_cf(ws)
    # ConditionalFormatting is keyed by identity/sqref in the internal dict;
    # mutating .sqref in place corrupts that dict, so pop/mutate/reinsert.
    rules = ws.conditional_formatting._cf_rules.pop(revel_cf)
    revel_cf.sqref.add(CellRange(new_range))
    ws.conditional_formatting._cf_rules[revel_cf] = rules

    wb.save(XLSX_PATH)
    print(f"Saved {XLSX_PATH}; new columns range {new_range}")

    for model_name in MODELS:
        counts = pd.Series(list(tiers_by_model[model_name].values())).value_counts(dropna=False)
        print(model_name, dict(counts))


if __name__ == "__main__":
    main()
