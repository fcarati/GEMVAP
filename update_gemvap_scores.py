"""
Overwrite the existing "GEMVAP 1 SCORE" / "GEMVAP 2 S" / "GEMVAP 3 S" columns
in expert_missense.xlsx with the raw consensus scores from the CURRENTLY
FITTED GEMVAP_1/2/3 models (same scoring run as score_expert_missense.py's
GEMVAP {1,2,3} CLASS columns, so the SCORE and CLASS columns are consistent
with each other).
"""
import sys
from pathlib import Path

import pandas as pd
import openpyxl

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))

from gemvap_pipeline.model import load_fit_result, apply_consensus_score

XLSX_PATH = HERE / "data" / "raw" / "expert_missense.xlsx"
SHEET = "Expert_missenses_91"
RAW_TSV = HERE / "data" / "raw" / "FBN1_tableS1_allmissense_gnomad4.tsv"
OUTPUT_DIR = HERE / "output" / "gemvap_notebook_gnomad4"
DATA_LAST_ROW = 92  # header is row 1; 91 variants in rows 2-92; rows 95-96 are footer notes

MODELS = ["GEMVAP_1", "GEMVAP_2", "GEMVAP_3"]
SCORE_COL_HEADER = {
    "GEMVAP_1": "GEMVAP 1 SCORE",
    "GEMVAP_2": "GEMVAP 2 S",
    "GEMVAP_3": "GEMVAP 3 S",
}


def main():
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb[SHEET]

    header = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    cdna_col = header["cDNA"]

    cdnas = [ws.cell(row=r, column=cdna_col).value for r in range(2, DATA_LAST_ROW + 1)]
    cdnas_clean = [c.strip() if isinstance(c, str) else None for c in cdnas]
    valid_cdnas = [c for c in cdnas_clean if c and c.startswith("c.")]

    raw = pd.read_csv(RAW_TSV, sep="\t", low_memory=False)
    raw = raw.drop_duplicates(subset="cDNA", keep="first").set_index("cDNA")

    missing = [c for c in valid_cdnas if c not in raw.index]
    if missing:
        print(f"WARNING: {len(missing)} cDNA values not found in {RAW_TSV.name}: {missing}")

    matched = raw.reindex(valid_cdnas)

    scores_by_model = {}
    for model_name in MODELS:
        fit = load_fit_result(OUTPUT_DIR / f"{model_name}_fit.pkl")
        top_predictors = fit["top_predictors"]
        case_thresholds = fit["rbc"]["threshold"]["case"]
        scores = apply_consensus_score(matched, top_predictors, case_thresholds)
        scores_by_model[model_name] = dict(zip(valid_cdnas, scores))

    for model_name in MODELS:
        col_idx = header[SCORE_COL_HEADER[model_name]]
        for i, cdna in enumerate(cdnas_clean):
            r = i + 2
            if cdna is None or cdna not in raw.index:
                value = "NA"
            else:
                score = scores_by_model[model_name][cdna]
                value = int(score) if pd.notna(score) else "NA"
            ws.cell(row=r, column=col_idx, value=value)

    wb.save(XLSX_PATH)
    print(f"Saved {XLSX_PATH}")
    for model_name in MODELS:
        vals = pd.Series(list(scores_by_model[model_name].values()))
        print(model_name, "min", vals.min(), "max", vals.max(), "mean", round(vals.mean(), 2))


if __name__ == "__main__":
    main()
