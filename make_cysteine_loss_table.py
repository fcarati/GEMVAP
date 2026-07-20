"""
Build publication-style table figures of FBN1 missense variants from
expert_missense.xlsx, one per user-supplied ClinVar-ID group, following the
same matplotlib-table styling as run_supplementary_table_coredataset.py's
_save_table().
"""
import re
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import openpyxl
import pandas as pd

HERE = Path(__file__).resolve().parent
XLSX_PATH = HERE / "data" / "raw" / "expert_missense.xlsx"
SHEET = "Expert_missenses_91"
OUT_DIR = HERE / "figures"
OUT_DIR.mkdir(parents=True, exist_ok=True)

GROUPS = {
    "Cysteine_Loss": {
        "title": "FBN1 missense variants associated with cysteine loss",
        "ids": [
            1380036, 426140, 495662, 549001, 636640, 502478, 406288, 956400, 495563,
            519758, 549070, 264089, 915814, 431935, 549150, 263660, 547309, 495599,
            626882, 222604, 1325422, 549229, 632819, 42391, 638559, 263898, 547338,
            633211, 547340, 429425, 492830, 222610, 632813, 495644, 1098776, 626877,
            423498, 42436,
        ],
    },
    "Cysteine_Gain": {
        "title": "FBN1 missense variants associated with cysteine gain",
        "ids": [
            222600, 548999, 549013, 449440, 180355, 519783, 570737, 381609, 42443,
            16466, 549024, 495594, 495629, 42402, 200084, 36118, 161245, 547296,
            439709, 928903, 549169, 155793, 618119, 384344,
        ],
    },
    "Others": {
        "title": "Other FBN1 missense variants (non-cysteine)",
        "ids": [
            163486, 36133, 36060, 42334, 16451, 42367, 263832, 549475, 42325, 263789,
            200177, 549019, 549173, 42339, 549180, 200100, 200198, 495569, 495558,
            200022, 549232, 178034, 36034, 617874, 495598, 200041, 373598, 200064,
            2413174,
        ],
    },
    "Excl_VUS_and_CysLoss": {
        "title": "FBN1 missense variants excluding VUS and cysteine-loss variants",
        "ids": [
            495569, 163486, 222600, 549475, 36133, 548999, 200177, 549013, 16466,
            549019, 549024, 495558, 36060, 42325, 42334, 16451, 549173, 495594,
            42339, 549180, 263789, 200022, 449440, 549232, 180355, 42367, 519783,
            495629, 570737, 42402, 200084, 200100, 178034, 200198, 36118, 381609,
            263832, 42443,
        ],
    },
    "VUS": {
        "title": "FBN1 missense variants of uncertain significance",
        "ids": [
            161245, 547296, 36034, 439709, 928903, 617874, 549169, 495598, 200041,
            155793, 373598, 200064, 618119, 384344, 2413174,
        ],
    },
}

TIER_COLOR = {
    "pathogenic": "#FF9900",   # matches the Excel CF's actual "pathogenic" orange (FFFF9900)
    "benign": "#6FA8DC",       # more saturated blue than the theme-accent1 tint, for clearer contrast
    "ambiguous": "#BFBFBF",    # grey, distinct from the header row's lighter #dddddd
}
EXPERT_TIER = {
    "Pathogenic": "pathogenic",
    "Likely pathogenic": "pathogenic",
    "Benign": "benign",
    "Likely benign": "benign",
    "Uncertain significance": "ambiguous",
}

DISPLAY_COLS = [
    "cDNA", "Protein change", "hg_38", "ClinVar ID", "Expert classification",
    "REVEL", "AlphaMissense", "GEMVAP 1", "GEMVAP 2", "GEMVAP 3",
    "REVEL match", "AlphaMissense match", "GEMVAP 1 match", "GEMVAP 2 match", "GEMVAP 3 match",
]
TIER_COLS = [
    "Expert classification", "REVEL", "AlphaMissense",
    "GEMVAP 1", "GEMVAP 2", "GEMVAP 3",
]
# Franklin / GeneBe / Varsome store raw ACMG evidence codes ("PP3 Supporting",
# "BP4 Strong", "UnMet") rather than a tier word directly -- EVIDENCE_CODE_COLS
# marks which TIER_COLS members need that PP3/BP4/UnMet -> tier translation.
EVIDENCE_CODE_COLS = {"Franklin", "GeneBe", "Varsome"}
ALL_DISPLAY_COLS = [
    "cDNA", "Protein change", "hg_38", "ClinVar ID", "Expert classification",
    "REVEL", "AlphaMissense", "GEMVAP 1", "GEMVAP 2", "GEMVAP 3",
    "Franklin", "GeneBe", "Varsome",
    "REVEL match", "AlphaMissense match", "GEMVAP 1 match", "GEMVAP 2 match", "GEMVAP 3 match",
    "Franklin match", "GeneBe match", "Varsome match",
]
ALL_TIER_COLS = TIER_COLS + ["Franklin", "GeneBe", "Varsome"]
MATCH_COLS = [
    "REVEL match", "AlphaMissense match", "GEMVAP 1 match", "GEMVAP 2 match", "GEMVAP 3 match",
    "Franklin match", "GeneBe match", "Varsome match",
]
MATCH_COL_SOURCE = {  # display column -> header of the colour-only column in the workbook
    "REVEL match": "REVEL",
    "AlphaMissense match": "AlphaMissense",
    "GEMVAP 1 match": "GEMVAP 1",
    "GEMVAP 2 match": "GEMVAP 2",
    "GEMVAP 3 match": "GEMVAP 3",
}
# Franklin/GeneBe/Varsome have no pre-existing colour-only column in the
# workbook (unlike REVEL/AlphaMissense/GEMVAP), so their match colour is
# computed here from the same tier logic used to colour the classification
# cells, rather than read from a cell fill.
COMPUTED_MATCH_SOURCE = {
    "Franklin match": "Franklin",
    "GeneBe match": "GeneBe",
    "Varsome match": "Varsome",
}
CONCORDANCE_GREEN = "#00B050"
CONCORDANCE_YELLOW = "#FFC000"
CONCORDANCE_RED = "#FF0000"


def concordance_color(predicted_tier, expert_tier):
    if predicted_tier is None or expert_tier is None:
        return None
    if predicted_tier == expert_tier:
        return CONCORDANCE_GREEN
    if predicted_tier == "ambiguous" or expert_tier == "ambiguous":
        return CONCORDANCE_YELLOW
    return CONCORDANCE_RED


def _argb_to_hex(argb: str) -> str | None:
    if not argb or argb == "00000000":
        return None
    return f"#{argb[-6:]}"


def resolve_tier(col_name: str, raw_value):
    """Map a cell's raw value to one of TIER_COLOR's keys ("pathogenic" /
    "benign" / "ambiguous"). Most TIER_COLS already store the tier word
    directly; Expert classification and the ACMG evidence-code columns
    (Franklin/GeneBe/Varsome: "PP3 Supporting" / "BP4 Strong" / "UnMet")
    need translating first."""
    if raw_value is None or raw_value == "NO DATA":
        return None
    if col_name == "Expert classification":
        return EXPERT_TIER.get(raw_value)
    if col_name in EVIDENCE_CODE_COLS:
        if "PP3" in raw_value:
            return "pathogenic"
        if "BP4" in raw_value:
            return "benign"
        return "ambiguous"  # "UnMet"
    return raw_value


def load_sheet():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb[SHEET]
    header = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}

    rows = []
    for r in range(2, 93):
        cv = ws.cell(row=r, column=header["ClinVar ID"]).value
        if cv is None:
            continue
        protein = ws.cell(row=r, column=header["PROTEIN"]).value or ""
        m = re.search(r"([A-Za-z]{3})(\d+)([A-Za-z]{3})", protein)
        position = int(m.group(2)) if m else None
        hg38_chrom = ws.cell(row=r, column=header["GRCh38Chromosome"]).value
        hg38_loc = ws.cell(row=r, column=header["GRCh38Location"]).value
        row = {
            "Position": position,
            "cDNA": ws.cell(row=r, column=header["cDNA"]).value,
            "Protein change": protein.strip("()"),
            "hg_38": f"chr{hg38_chrom}:{hg38_loc}" if hg38_chrom and hg38_loc else "",
            "ClinVar ID": int(cv),
            "Expert classification": ws.cell(row=r, column=header["EXPERT CLASSIFICATION"]).value,
            "REVEL": ws.cell(row=r, column=header["REVEL P"]).value,
            "AlphaMissense": ws.cell(row=r, column=header["AlphaMissense P"]).value,
            "GEMVAP 1": ws.cell(row=r, column=header["GEMVAP 1 CLASS"]).value,
            "GEMVAP 2": ws.cell(row=r, column=header["GEMVAP 2 CLASS"]).value,
            "GEMVAP 3": ws.cell(row=r, column=header["GEMVAP 3 CLASS"]).value,
            "Franklin": ws.cell(row=r, column=header["Franklin PP3/BP4"]).value,
            "GeneBe": ws.cell(row=r, column=header["GeneBe"]).value,
            "Varsome": ws.cell(row=r, column=header["Varsome"]).value,
        }
        expert_tier = resolve_tier("Expert classification", row["Expert classification"])
        for display_col, source_header in MATCH_COL_SOURCE.items():
            cell = ws.cell(row=r, column=header[source_header])
            row[display_col] = ""  # colour-only column, no text (matches the workbook)
            fill_hex = _argb_to_hex(cell.fill.fgColor.rgb)
            if fill_hex is None:
                # Workbook left this cell unfilled (e.g. REVEL has no raw score for a
                # start-loss variant) even though a curated tier still exists -- fall
                # back to computing concordance from that tier instead of leaving blank.
                predicted_tier = resolve_tier(source_header, row[source_header])
                fill_hex = concordance_color(predicted_tier, expert_tier)
            row[f"{display_col}__hex"] = fill_hex

        for display_col, source_col in COMPUTED_MATCH_SOURCE.items():
            predicted_tier = resolve_tier(source_col, row[source_col])
            row[display_col] = ""  # colour-only column, no text
            row[f"{display_col}__hex"] = concordance_color(predicted_tier, expert_tier)
        rows.append(row)
    return pd.DataFrame(rows)


def make_table(df_all: pd.DataFrame, group_name: str, title: str, ids: list,
               display_cols: list = DISPLAY_COLS, tier_cols: list = TIER_COLS):
    wanted = set(ids)
    df = df_all[df_all["ClinVar ID"].isin(wanted)].sort_values("Position").reset_index(drop=True)

    found_ids = set(df["ClinVar ID"])
    missing = wanted - found_ids
    if missing:
        print(f"[{group_name}] WARNING: {len(missing)} ClinVar IDs not found in the sheet: {sorted(missing)}")
    print(f"[{group_name}] {len(df)} variants matched")

    csv_path = OUT_DIR / f"FBN1_{group_name}_Variants.csv"
    hex_cols = [c for c in df.columns if c.endswith("__hex")]
    df.drop(columns=["Position"] + hex_cols).to_csv(csv_path, index=False)
    print(f"[{group_name}] Saved {csv_path}")

    cell_text = df[display_cols].astype(str).values.tolist()

    # Precise sizing: matplotlib's table fills the axes bbox exactly at scale=1,
    # so we control row height (and thus overall table height) via the AXES
    # height in inches, and reserve separate fixed-inch margins above/below for
    # the title and footnote -- avoids the row-overflow that .scale(y>1) causes
    # when it inflates rows beyond whatever axes bbox height was guessed.
    fontsize = 8.5
    row_height_in = fontsize / 72 * 1.6
    n_total_rows = len(df) + 1  # + header
    table_height_in = row_height_in * n_total_rows
    title_margin_in = 0.55
    footnote_margin_in = 0.35
    fig_width = 12 + 0.55 * len(display_cols)
    fig_height = table_height_in + title_margin_in + footnote_margin_in

    fig = plt.figure(figsize=(fig_width, fig_height))
    ax = fig.add_axes([0.01, footnote_margin_in / fig_height, 0.98, table_height_in / fig_height])
    ax.axis("off")

    tbl = ax.table(cellText=cell_text, colLabels=display_cols, cellLoc="center", loc="center")
    tbl.auto_set_font_size(False)
    tbl.set_fontsize(fontsize)
    tbl.auto_set_column_width(col=list(range(len(display_cols))))

    for (row, col), cell in tbl.get_celld().items():
        if row == 0:
            cell.set_text_props(weight="bold", color="white")
            cell.set_facecolor("black")
            continue
        col_name = display_cols[col]
        if col_name in tier_cols:
            raw_value = df.iloc[row - 1][col_name]
            tier = resolve_tier(col_name, raw_value)
            color = TIER_COLOR.get(tier)
            if color:
                cell.set_facecolor(color)
        elif col_name in MATCH_COLS:
            color = df.iloc[row - 1][f"{col_name}__hex"]
            if isinstance(color, str):
                cell.set_facecolor(color)

    fig.text(
        0.5, 1 - title_margin_in * 0.45 / fig_height,
        f"{title} (n={len(df)})", ha="center", va="center", fontsize=13,
    )
    fig.text(
        0.5, footnote_margin_in * 0.45 / fig_height,
        "Classification columns: orange = pathogenic-tier, blue = benign-tier, grey = ambiguous/uncertain.  "
        "\"match\" columns: green = GEMVAP agrees with expert classification, yellow = disagrees but one call was "
        "ambiguous, red = disagrees and neither was ambiguous.",
        ha="center", va="center", fontsize=8, style="italic",
    )

    png_path = OUT_DIR / f"FBN1_{group_name}_Variants.png"
    fig.savefig(png_path, dpi=300, bbox_inches="tight")
    plt.close(fig)
    print(f"[{group_name}] Saved {png_path}")
    return png_path


def main():
    df_all = load_sheet()
    for group_name, spec in GROUPS.items():
        make_table(df_all, group_name, spec["title"], spec["ids"])

    make_table(
        df_all, "All_Variants", "All FBN1 missense variants",
        df_all["ClinVar ID"].tolist(),
        display_cols=ALL_DISPLAY_COLS, tier_cols=ALL_TIER_COLS,
    )


if __name__ == "__main__":
    main()
